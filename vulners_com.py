# 'EStroev'
from zipfile import ZipFile
import re
import os
import ijson
from datetime import datetime
import argparse
import csv
import openpyxl


def file_writer(outPathFile, data):
    with open(outPathFile, 'a', encoding='utf-8') as outFile:
        for id in data:
            outFile.write(
                f"ID: {data[id]['id']}\n\t"
                f"Title: {data[id]['title']}\n\t"
                f"URL: {data[id]['url']}\n\t"
                f"CVE: {'; '.join(data[id]['cve'])}\n\t"
                f"CVSS: {data[id]['cvss']}\n\t"
                f"Vector: {data[id]['vector']}\n\t"
                f"Type: {data[id]['type']}\n\n"
            )
    print(f'[+] Write {len(data)} entries to {outPathFile}')


def csv_writer(outPathFile, data):
    writeData = dict()
    with open(outPathFile, 'a', newline='') as csv_out:
        csv_out_writer = csv.writer(csv_out, delimiter=',')
        for id in data:
            csv_out_writer.writerow(
                [
                    data[id]['cve'],
                    data[id]['cvss'],
                    data[id]['vector'],
                    f"{data[id]['id']}:{data[id]['url']}",
                    data[id]['ip']
                ]
            )
        print(f'[+] Write {len(data)} entries to {outPathFile}')


def xslx_writer(outPath, title, data):
    writeList1 = dict()
    for cve in data:
        writeList1[cve] = {
            'cvss': data[cve]['cvss'],
            'vector': data[cve]['vector'],
            'url': set(data[cve]['url']) if data[cve]['url'] else None,
            'ip': data[cve]['ip']
        }


    writeList = list()
    for item in writeList1:
        writeList.append(
            [
                item,
                writeList1[item]['cvss'],
                writeList1[item]['vector'],
                '\n'.join(writeList1[item]['url']) if writeList1[item]['url'] else 'Not found',
                writeList1[item]['ip']
            ]
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    for j in range(len(writeList[0])):
        ws.cell(row=1, column=j+1).value = writeList[0][j]

    for i in range(len(writeList[1:])):
        for j in range(len(writeList[0])):
            ws.cell(row=i+2, column=j+1).value = writeList[i+1][j]

    wb.save(outPath)
    print('[+] Write "{}" ({} entries) to {}'.format(title, len(writeList), outPath))


def exploit_searcher(inPath, outPathFile, cve_file, inputData, searchObject):
    data = dict()

    zFile = ZipFile(cve_file, 'r')
    for finfo in zFile.infolist():
        iFile = zFile.open(finfo)

        print('[+] Open %s ' % cve_file)
        parser = ijson.parse(iFile)

        id, title, url, type, descr, cve, cvss, vector = None, None, None, None, None, [], None, None
        searchDict = {
            'id': '',
            'title': '',
            'url': '',
            'description': '',
            'cve': '',
            'cvss': '',
            'vector': ''
        }
        for prefix, event, value in parser:
            if prefix == '_index._index' == 'bulletins':
                id, title, url, type, descr, cve, cvss, vector = None, None, None, None, None, [], None, None
                searchDict = {
                    'id': '',
                    'title': '',
                    'url': '',
                    'description': '',
                    'cve': '',
                    'cvss': '',
                    'vector': ''
                }
            elif prefix == 'item._id':
                id = value
                searchDict['id'] = id
            elif prefix == 'item._source.title':
                title = value
                searchDict['title'] = title
            elif prefix == 'item._source.type':
                type = value
            elif prefix == 'item._source.href':
                url = value
                searchDict['url'] = url
            elif prefix == 'item._source.description':
                descr = value
                searchDict['description'] = descr
            elif prefix == 'item._source.cvss.score':
                cvss = str(value)
                searchDict['cvss'] = cvss
            elif prefix == 'item._source.cvss.vector':
                vector = value
                searchDict['vector'] = vector
            elif prefix == 'item._source.cvelist.item':
                cve.append(value)
                searchDict['cve'] = cve
            if id and title and url and type and descr and cvss:
                for line in inputData:
                    if line:
                        ip, pattern = line.split(';')
                        if pattern == id:
                            print(f'[+] [{ip}] Find {pattern} in {cve_file}: {id} - {title}')
                            data[pattern] = {
                                'ip': ip,
                                'cvss': f'={cvss}',
                                'vector': vector,
                                'url': list()
                            }
                id, title, url, descr, type, cve, cvss, vector = None, None, None, None, None, [], None, None
                continue
            else:
                continue

    for root, dirs, files in os.walk(inPath):
        for file in files:
            if file.endswith('.zip'):
                inFilePath = os.path.join(root, file)
                zFile = ZipFile(inFilePath, 'r')
                for finfo in zFile.infolist():
                    iFile = zFile.open(finfo)

                    print('[+] Open %s ' % inFilePath)
                    parser = ijson.parse(iFile)

                    id, title, url, type, descr, cve, cvss, vector = None, None, None, None, None, [], None, None
                    searchDict = {
                        'id': '',
                        'title': '',
                        'url': '',
                        'description': '',
                        'cve': '',
                        'cvss': '',
                        'vector': ''
                    }
                    for prefix, event, value in parser:
                        if prefix == '_index._index' == 'bulletins':
                            id, title, url, type, descr, cve, cvss, vector = None, None, None, None, None, [], None, None
                            searchDict = {
                                'id': '',
                                'title': '',
                                'url': '',
                                'description': '',
                                'cve': '',
                                'cvss': '',
                                'vector': ''
                            }
                        elif prefix == 'item._id':
                            id = value
                            searchDict['id'] = id
                        elif prefix == 'item._source.title':
                            title = value
                            searchDict['title'] = title
                        elif prefix == 'item._source.type':
                            type = value
                        elif prefix == 'item._source.href':
                            url = value
                            searchDict['url'] = url
                        elif prefix == 'item._source.description':
                            descr = value
                            searchDict['description'] = descr
                        elif prefix == 'item._source.cvss.score':
                            cvss = str(value)
                            searchDict['cvss'] = cvss
                        elif prefix == 'item._source.cvss.vector':
                            vector = value
                            searchDict['vector'] = vector
                        elif prefix == 'item._source.cvelist.item':
                            cve.append(value)
                            searchDict['cve'] = cve

                        if id and title and url and type and descr and cvss:
                            for line in inputData:
                                if line:
                                    ip, pattern = line.split(';')
                                    if pattern in searchDict[searchObject]:
                                        print(f'[+] [{ip}] Find {pattern} in {file}: {id} - {title}')
                                        if pattern in data:
                                            if '{id}:{url}' not in data[pattern]['url']:
                                                data[pattern]['url'].append(f'{id}:{url}')
                                            data[pattern]['ip'] = ip
                                            data[pattern]['id'] = id
                                            # data[pattern]['url'] = [url]
                            id, title, url, descr, type, cve, cvss, vector = None, None, None, None, None, [], None, None
                            continue
                        else:
                            continue

    if data:
        for line in inputData:
            if line:
                ip, pattern = line.split(';')
                if pattern not in data:
                    data[pattern] = {
                        'ip': ip,
                        'id': None,
                        'title': None,
                        'url': None,
                        'type': None,
                        'cve': None,
                        'cvss': None,
                        'vector': None,

                    }
        xslx_writer(outPathFile, 'test', data)


def worker(inPath, outPathFile, pattern, searchObject):

    for root, dirs, files in os.walk(inPath):
        for file in files:
            if file.endswith('.zip'):
                inFilePath = os.path.join(root, file)
                zFile = ZipFile(inFilePath, 'r')
                for finfo in zFile.infolist():
                    iFile = zFile.open(finfo)

                    print('[+] Open %s ' % inFilePath)
                    parser = ijson.parse(iFile)

                    data = dict()

                    id, title, url, type, descr, cve, cvss, vector = None, None, None, None, None, [], None, None
                    searchDict = {
                        'id': '',
                        'title': '',
                        'url': '',
                        'description': '',
                        'cve': '',
                        'cvss': '',
                        'vector': ''
                    }
                    for prefix, event, value in parser:
                        if prefix == '_index._index' == 'bulletins':
                            id, title, url, type, descr, cve, cvss, vector = None, None, None, None, None, [], None, None
                            searchDict = {
                                'id': '',
                                'title': '',
                                'url': '',
                                'description': '',
                                'cve': '',
                                'cvss': '',
                                'vector': ''
                            }
                        elif prefix == 'item._id':
                            id = value
                            searchDict['id'] = id
                        elif prefix == 'item._source.title':
                            title = value
                            searchDict['title'] = title
                        elif prefix == 'item._source.type':
                            type = value
                        elif prefix == 'item._source.href':
                            url = value
                            searchDict['url'] = url
                        elif prefix == 'item._source.description':
                            descr = value
                            searchDict['description'] = descr
                        elif prefix == 'item._source.cvss.score':
                            cvss = str(value)
                            searchDict['cvss'] = cvss
                        elif prefix == 'item._source.cvss.vector':
                            vector = value
                            searchDict['vector'] = vector
                        elif prefix == 'item._source.cvelist.item':
                            cve.append(value)
                            searchDict['cve'] = cve

                        if id and title and url and type and descr and cvss:
                            if pattern.search(searchDict[searchObject]):
                                print(f'[+] Find in {searchObject}: {searchDict[searchObject]}')
                                data[id] = {
                                    'id': id,
                                    'title': title,
                                    'url': url,
                                    'type': type,
                                    'cve': pattern,
                                    'cvss': cvss,
                                    'vector': vector
                                }
                            id, title, url, descr, type, cve, cvss, vector = None, None, None, None, None, [], None, None
                            continue
                        else:
                            continue

                    if data:
                        file_writer(outPathFile, data)


def main():
    parser = argparse.ArgumentParser(description='Vulners.com DB parser')
    parser.add_argument('-o', dest='outFolder', action='store', help='Output folder')
    parser.add_argument('-f', dest='inPath', action='store', help='Input path')
    parser.add_argument('-p', dest='pattern', action='store', help='Searching pattern')
    parser.add_argument('-d', dest='searchObject', action='store', default='cve', help='Search object')
    parser.add_argument('-i', dest='inputFile', action='store', help='Input file')
    parser.add_argument('-c', dest='cveFile', action='store', help='CVE DB Input file')

    args = parser.parse_args()
    if not args.inPath:
        print('[-] You must specify an existing input path!')
        exit(-1)
    elif not os.path.exists(args.inPath):
        print('[-] Input path %s does not exist!' % os.path.abspath(args.inPath))
        exit(-1)
    # if not args.pattern:
    #     print('[-] You must specify a searching pattern!')
    #     exit(-1)
    if not args.searchObject:
        print('[-] You must specify a searching object (id, title, url, description, cve, cvss)!')
        exit(-1)
    elif args.searchObject not in ['id', 'title', 'url', 'description', 'cve', 'cvss']:
        print('[-] You must specify one of the following objects: id, title, url, description, cve, cvss!')
        exit(-1)
    if not args.outFolder:
        print('[-] You must specify an existing path to the output folder!')
        exit(-1)
    elif not os.path.exists(args.outFolder):
        print(f'[-] Output folder {os.path.abspath(args.outFolder)} does not exist!')
        os.makedirs(args.outFolder)
        print(f'[+] Create output folder {args.outFolder}')

    # outFile = os.path.join(args.outFolder, '{}_{}.txt'.format(args.searchObject, args.pattern.replace("\s", "_")))
    outFile = os.path.join(args.outFolder, '{}_{}.xlsx'.format(args.searchObject, args.inputFile))
    startTime = datetime.now()
    print(startTime.strftime('[*] Start time: %d.%m.%Y %H:%M:%S'))

    # pattern = re.compile(args.pattern)
    # pattern = args.pattern
    if os.path.exists(outFile):
        os.remove(outFile)
        print(f'[-] Output file {outFile} exist! Removed it!')
    else:
        with open(outFile, 'w') as tmpFile:
            pass

    # worker(args.inPath, outFile, pattern, args.searchObject)
    with open(args.inputFile) as inF:
        inputData = inF.read().split('\n')

    exploit_searcher(args.inPath, outFile, args.cveFile, inputData, args.searchObject)

    endTime = datetime.now()
    print('[*] Total elapsed time - {0} seconds'.format((endTime - startTime).seconds))
    print(endTime.strftime('[*] End time: %d.%m.%Y %H:%M:%S'))

if __name__ == '__main__':
    main()